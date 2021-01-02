import io
import os

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Protection
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


class XlsReportBuilder:
    """
    This class contains all the module that helps to build the xls report.
    """
    MODENEW = "newworkbook"
    MODEREAD = "read"
    MODEWRITE = "write"
    MODEAPPEND = "append"

    def __init__(self, file_loc, mode=MODEREAD):
        if (mode == self.MODEREAD) and (not os.path.isfile(file_loc)):
            raise RuntimeError("Failed to open xls file [{}]".format(file_loc))
        self.mode, self.file_loc = mode, file_loc
        self.workbook, self.sheets, self.sheet = None, None, None

    def __enter__(self):
        self.open_work_book()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close_work_book()

    def open_work_book(self, component_type=None, result_file=None, index=None):
        if self.mode == self.MODENEW:
            self.workbook = Workbook(self.file_loc)
            self.workbook.save(result_file)
        elif self.mode == self.MODEWRITE:
            self.workbook = openpyxl.load_workbook(self.file_loc)
            self.sheet = self.workbook.create_sheet(component_type, index - 1)
            for sobj in range(len(self.workbook.sheetnames)):
                if self.workbook.sheetnames[sobj] == component_type:
                    break
            self.workbook.active = sobj
            self.sheet = self.workbook.active
            self.workbook.save("{}".format(result_file))

        elif self.mode == self.MODEAPPEND:
            self.workbook = openpyxl.load_workbook(self.file_loc)
            for s in range(len(self.workbook.sheetnames)):
                if self.workbook.sheetnames[s] == component_type:
                    break
            self.workbook.active = s
            self.sheet = self.workbook.active

    def close_work_book(self):
        if self.mode == self.MODEREAD:
            pass
        if self.mode == self.MODEWRITE:
            with io.open(self.file_loc, 'wb') as file:
                self.workbook.save(file)

    def write_cell(self, row_no, col_no, cell_value):
        self.sheet.cell(row=row_no, column=col_no).value = cell_value

    @staticmethod
    def set_style(sheet, xls_style):
        """
        This method is used to set the styles on all mutable field.
        :param sheet:
        :param xls_style:
        :return:
        """
        for style in xls_style:
            if style == "header_font":
                sheet.font = xls_style[style]
            if style == "border":
                sheet.border = xls_style[style]
            if style == "alignment":
                sheet.alignment = xls_style[style]
            if style == "data_field_font":
                sheet.font = xls_style[style]
            if style == "background":
                sheet.fill = xls_style[style]

    def write_row(self, row_no, from_column, value_list, xlsstyle, column_width):
        """
        This method is used to write the content in the row.
        :param row_no:
        :param from_column:
        :param value_list:
        :param xlsstyle:
        :param column_width:
        :return:
        """
        col_no = from_column
        if type(value_list) == list:
            for val in value_list:
                self.sheet.cell(row=row_no, column=col_no).value = val
                sheet_obj = self.sheet.cell(row=row_no, column=col_no)
                XlsReportBuilder.set_style(sheet_obj, xlsstyle)
                col_no += 1
        else:
            self.sheet.cell(row=row_no, column=col_no).value = value_list
        self.column_width(value_list, column_width)

    def write_column(self, from_row, col_no, value_list):
        """
        This method is used to write the content in the column
        :param from_row:
        :param col_no:
        :param value_list:
        :return:
        """
        row_no = from_row
        for val in value_list:
            self.sheet.cell(row=row_no, column=col_no).value = val
            row_no += 1

    @staticmethod
    def xls_sheet_style(request_type):
        """
        This method is used to set the styles on all mutable field.
        :param request_type:
        :return:
        """
        header_font = Font(name='Calibri', size=12, bold=True, italic=False, vertAlign=None,
                           underline='none', strike=False, color='00000000')
        data_field_font = Font(name='Calibri', size=10, bold=False, italic=False, vertAlign=None,
                               underline='none', strike=False, color='00000000')
        background = PatternFill(start_color="0000FFFF", end_color="0000FFFF", fill_type="solid")
        border = Border(left=Side(border_style=None, color='00000000'),
                        right=Side(border_style=None, color='00000000'),
                        top=Side(border_style=None, color='00000000'),
                        bottom=Side(border_style=None, color='00000000'),
                        diagonal=Side(border_style=None, color='00000000'),
                        diagonal_direction=0,
                        outline=Side(border_style=None, color='00000000'),
                        vertical=Side(border_style=None, color='00000000'),
                        horizontal=Side(border_style=None, color='00000000'))
        alignment = Alignment(horizontal='center', vertical='center', text_rotation=0,
                              wrap_text=True, shrink_to_fit=True, indent=0)
        protection = Protection(locked=False, hidden=False)

        header_style = {"header_font": header_font, "border": border,
                        "alignment": alignment, "protection": protection, "background": background}
        data_field_style = {"data_field_font": data_field_font, "border": border,
                            "alignment": alignment, "protection": protection}

        if request_type == "header":
            return header_style
        else:
            return data_field_style

    def column_width(self, data, changed_index_list):
        """
        This method is used to increase the cells width based on the default value as well as
         changed index value.
        :param data:
        :param changed_index_list:
        :return:
        """
        column_widths = []
        for row in data:
            default_width = 20
            extended_width = 90
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell) + default_width
                else:
                    if i in changed_index_list:
                        column_widths += [len(cell) + extended_width]
                    else:
                        column_widths += [len(cell) + default_width]
        for i, column_width in enumerate(column_widths):
            self.sheet.column_dimensions[get_column_letter(i + 1)].width = column_width

    @staticmethod
    def cell_alignment(result_file=None, index_list=None):
        """
        This method is used to wrap the text of all the fields.
        :param result_file:
        :param index_list:
        :return:
        """
        wb = openpyxl.load_workbook(result_file)
        index_list = [x + 1 for x in index_list]
        ws = wb.active
        for rows in ws.iter_rows(min_row=2, max_row=None, min_col=None, max_col=None):
            for cell in rows:
                if cell.col_idx in index_list:
                    cell.alignment = cell.alignment.copy(wrapText=True, horizontal='left',
                                                         vertical='justify')
                else:
                    cell.alignment = cell.alignment.copy(wrapText=True, horizontal='center',
                                                         vertical='justify')
        wb.save(result_file)
